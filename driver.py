import course as c
import globals as g
from openpyxl.reader.excel import load_workbook

"""---------------------------------------
Creates course objects from a text file fileName

File format:
    Course numbers are in column:
        A for summer classes
        F for Autumn
        K for winter
        P for Spring

"""
def buildClasses(fileName=g.COURSE_FILE):
    wb = load_workbook(fileName)
    ws = wb.active
    #Build all summer courses (starting in column A)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
        if ws["A"+str(row)] is not None:
            num = ws["A"+str(row)].value
            time = ws["B"+str(row)].value
            day = ws["C"+str(row)].value
            cap = ws["D"+str(row)].value
            result = c.Course(0, num,time,day,cap)
            if result.num != -1:
                g.SUMCAT.append(result)

    #Build all Autumn courses (starting in column F)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
        if ws["F"+str(row)] is not None:
            num = ws["F"+str(row)].value
            time = ws["G"+str(row)].value
            day = ws["H"+str(row)].value
            cap = ws["I"+str(row)].value
            result = c.Course(0, num,time,day,cap)
            if result.num != -1:
                g.AUTCAT.append(result)
    #Build all Winter courses (starting in column K)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
        if ws["K"+str(row)] is not None:
            num = ws["K"+str(row)].value
            time = ws["L"+str(row)].value
            day = ws["M"+str(row)].value
            cap = ws["N"+str(row)].value
            result = c.Course(0, num,time,day,cap)
            if result.num != -1:
                g.WINCAT.append(result)
    #Build all Spring courses (starting in column P)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
        if ws["P"+str(row)] is not None:
            num = ws["P"+str(row)].value
            time = ws["Q"+str(row)].value
            day = ws["R"+str(row)].value
            cap = ws["S"+str(row)].value
            result = c.Course(0, num,time,day,cap)
            if result.num != -1:
                g.SPRCAT.append(result)


buildClasses()
print("SUMMER CATALOG: ")
print("----------------------------------------")

for i in range(len(g.SUMCAT)):
    g.SUMCAT[i].display()
print()
print("AUT CATALOG: ")
print("----------------------------------------")
for i in range(len(g.AUTCAT)):
    g.AUTCAT[i].display()
print()
print("WIN CATALOG")
print("----------------------------------------")

for i in range(len(g.WINCAT)):
    g.WINCAT[i].display()
print()
print("SPR CATALOG")
print("----------------------------------------")
for i in range(len(g.SPRCAT)):
    g.SPRCAT[i].display()
print()