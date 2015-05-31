"""
This file provides the method to build the faculty.
"""
from faculty import faculty
import globals as g
from openpyxl.reader.excel import load_workbook

"""---------------------------------------
Creates faculty objects from a text file fileName
Pre:
    Using a file formatted exactly like list_of_faculty.xlsx
    No professor classes have been made
Post:
    Faculty classes have been generated with relevant data
File format:
    Column A = Last name
    Column B = Full-time status
    Column C = Number of classes
    Column D = Number of students advised
    Column E = Expertise 1
    Column F = Expertise 2 (if available)
    Column G = Expertise 3 (if available)
"""
def buildFaculty(fileName=g.FACULTY_FILE):
    
    wb = load_workbook(fileName)
    ws = wb.active
    
    # Build faculty from first readable row to last
    rowNum = 5
    while ws.cell(row = rowNum, column = 1).value is not None:  # Rows >= 5 in the file have relevant data
        lastName = ws.cell(row = rowNum, column = 1).value
        
        if ws.cell(row = rowNum, column = 2).value == 'Y':
            fullTime = 1
        else:
            fullTime = 0
            
        numClasses = ws.cell(row = rowNum, column = 3).value + 0.
        studentsAdvised = ws.cell(row = rowNum, column = 4).value
        
        expertise = [ws.cell(row = rowNum, column = 5).value]

        if ws.cell(row = rowNum, column = 6).value != None:
            expertise.append(ws.cell(row = rowNum, column = 6).value)
        if ws.cell(row = rowNum, column = 7).value != None:
            expertise.append(ws.cell(row = rowNum, column = 7).value)
        if ws.cell(row = rowNum, column = 6).value != None:
            expertise.append(ws.cell(row = rowNum, column = 6).value)
        if ws.cell(row = rowNum, column = 7).value != None:
            expertise.append(ws.cell(row = rowNum, column = 7).value)
            
        g.allFaculty.append(faculty(lastName, fullTime, numClasses, studentsAdvised, expertise))
        
        rowNum += 1
        
    g.ADVISED_STUDENTS_PER_QUARTER = 0
    for i in g.allFaculty:
        g.ADVISED_STUDENTS_PER_QUARTER += i.studentsAdvised
