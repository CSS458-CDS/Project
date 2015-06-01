"""
This file provides the method to build course catalogs.
"""
import course as c
import globals as g
from openpyxl.reader.excel import load_workbook
import time

"""---------------------------------------
Creates course objects from a text file fileName

Pre:
    Using a file formatted exactly like list_of_classes.xlsx
    Global variables WINCAT, AUTCAT, SPRCAT, SUMCAT are all empty lists
Post:
    WINCAT, AUTCAT, SPRCAT, SUMCAT contain course objects will all relevant data from the file

File format:
    Course numbers are in column:
        A for summer classes
        F for Autumn
        K for winter
        P for Spring

    Every course has a 3 number identify
    Sometimes there's a section, sometimes there's a prefix


"""
def buildClasses(fileName=g.COURSE_FILE):
    wb = load_workbook(fileName)
    ws = wb.active
    #Build all summer courses (starting in column A)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
        if ws["A"+str(row)] is not None:
            result = make_one_course(ws, 'A', row)
            if result is not None:
                if result.num != -1:
                    result.getCourseData()
                    result.getTitle()
                    result.set_days()
                    if result.title != "Default Title":
                        g.SUMCAT.append(result)

    #Build all Autumn courses (starting in column F)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
            if ws["F"+str(row)] is not None:
                result = make_one_course(ws, 'F', row)
            if result is not None:
                if result.num != -1:
                    result.getCourseData()
                    result.getTitle()
                    result.set_days()
                    if result.title != "Default Title":
                        g.AUTCAT.append(result)

    #Build all Winter courses (starting in column K)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
        if ws["K"+str(row)] is not None:
            result = make_one_course(ws, "K", row)
            if result is not None:
                if result.num != -1:
                    result.getCourseData()
                    result.getTitle()
                    result.set_days()
                    if result.title != "Default Title":
                        g.WINCAT.append(result)
    #Build all Spring courses (starting in column P)
    for row in range(g.COURSE_FILE_FIRST,g.COURSE_FILE_LAST):
        if ws["P"+str(row)] is not None:
            result = make_one_course(ws, 'P', row)
            if result is not None:
                if result.num != -1:
                    result.getCourseData()
                    result.getTitle()
                    result.set_days()
                    if result.title != "Default Title":
                        g.SPRCAT.append(result)


def make_one_course(ws, startColumn, startRow):
    if ws[startColumn+str(startRow)] is None:
        return None

    row = startRow
    numCol = startColumn
    timeCol = chr(ord(startColumn)+1)
    dayCol = chr(ord(startColumn)+2)
    capCol = chr(ord(startColumn)+3)
    num = ws[numCol+str(row)].value
    #Handle capstone
    if num == 497:
        result = c.Course(courseNum=num, time="-1", day="-1", cap=g.ADVISED_STUDENTS_PER_QUARTER)
        result.taughtBy = -1
        result.dept = "CSS"
        return result

    timestr = str(ws[timeCol+str(row)].value)
    if "-" in timestr:
        timestr = timestr[0:timestr.index("-")]
    while timestr[0] == "0":
        timestr = timestr[1:]
    if len(timestr) == 1:
        timestr += ":00"
    if len(timestr) > 5 and timestr.index(":") == 1:
        timestr = timestr[0:4]
    elif len(timestr) > 5:
        timestr = timestr[0:5]
    while len(timestr) < 5:
        timestr = " " + timestr
    day = ws[dayCol+str(row)].value
    cap = ws[capCol+str(row)].value
    result = c.Course(0, num,timestr,day,cap)
    #Nicely behaving course number
    if type(result.num) == int:
        result.dept = "CSS"
        return result
    #Course number with anomalies (csskl, bcusp and the like)
    else:
        #Has parens for some reason
        if ")" in result.num:
            noparens = result.num[0:result.num.index("(")]
            result.num = noparens.strip()
        #Special case: Ends with a letter (section)
        if 65 <= ord(result.num[len(result.num) - 1]) <= 90:
            #Take the last char as the section
            result.section = result.num[len(result.num) - 1]
            #Remove the char from the number string
            result.num = result.num[0:(len(result.num)-1)]


        #Result just had a section, not a department
        if len(result.num) == 3:
            result.num = int(result.num)
            result.dept = "CSS"
            return result
        deptstr = result.num[0:3]
        if deptstr == "SKL":
            result.dept = "CSSSKL"
            result.num = int(result.num[3:len(result.num)])
            return result
        deptstr = result.num[0:5]
        if deptstr == "BCUSP":
            result.dept = "BCUSP"
            result.num = int(result.num[5:len(result.num)])
            return result
#       If it isnt skl or bcsup, it's css
        result.dept = "CSS"
        stripped = result.num.strip()
        if len(stripped) == 3:
            result.num = int(stripped)
            return result
        #Case: Lab
        if "Lab" in result.num:
            result.num = result.num[0:3]
            result.lab = True
            return result
        #Case: Women in stem is offered as 2 different courses, but is one course
        if result.num == "205/405":
            result.num = 405
            return result
        print("Error reading course: ", end="", sep="")
        result.display()

def print_catalog(cat):
    for i in range(len(cat)):
        cat[i].display()
"""
Pre: AUTCAT, SPRCAT, SUMCAT, WINCAT are all populated
Post: A list containing all the electives from the parameter catalog is returned
"""
def get_electives_from(catalog):
    retval = []
    for i in range(len(catalog)):
        if catalog[i].elective:
            retval.append(catalog[i])
    return retval


def print_catalogs():
    if len(g.AUTCAT) == 0:
        g.initialize_globals()

    print("Summer Catalog: ")
    for i in g.SUMCAT:
        i.display(title=True)
    print("-------------------------------------------------------------------")

    print("Autumn Catalog")
    for i in g.AUTCAT:
        i.display(title=True)
    print("-------------------------------------------------------------------")

    print("Winter Catalog")
    for i in g.WINCAT:
        i.display(title=True)
    print("-------------------------------------------------------------------")
    print("Spring Catalog")
    for i in g.SPRCAT:
        i.display(title=True)
    print("-------------------------------------------------------------------")
#Test that the catalogs were built correctly
#print_catalogs()
